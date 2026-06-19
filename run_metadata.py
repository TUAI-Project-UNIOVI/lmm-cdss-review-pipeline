"""Stage 1 — metadata retrieval and corpus assembly.

Orchestrates two substeps:
  1. Fetch  — pull raw metadata from each database → outputs/metadata/*_raw.csv
  2. Corpus — merge raws, deduplicate, normalise   → outputs/metadata/corpus.*

Substep modules live in metadata/fetch/ and metadata/corpus/.

Usage:
    python run_metadata.py                        # run both substeps
    python run_metadata.py --fetch-only           # fetch raws only
    python run_metadata.py --corpus-only          # build corpus from existing raws
    python run_metadata.py --sources pubmed ieee  # fetch selected sources, then build corpus

Options:
    --fetch-only     Stop after saving *_raw.csv files; do not build corpus.
    --corpus-only    Skip fetch; build corpus from whichever *_raw.csv files exist.
    --sources        Space-separated subset of sources to fetch: pubmed ieee wos (default: all).
    --max N          Override MAX_RESULTS from config.py.
"""

import argparse
import logging
import os
import sys
from pathlib import Path

import joblib
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font

import config
from metadata.fetch.pubmed import PubMedFetcher
from metadata.fetch.ieee import IEEEExportLoader
from metadata.fetch.wos import WoSExportLoader
from metadata.preprocessing.dedup import deduplicate_corpus, build_duplicate_map
from utils import setup_logging, ensure_output_dir

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Stage 1: metadata retrieval and corpus assembly")
    mode = p.add_mutually_exclusive_group()
    mode.add_argument("--fetch-only", action="store_true",
                      help="Fetch raw metadata only; skip corpus assembly.")
    mode.add_argument("--corpus-only", action="store_true",
                      help="Build corpus from existing raw files; skip fetch.")
    p.add_argument(
        "--sources", nargs="+",
        choices=["pubmed", "ieee", "wos"],
        default=["pubmed", "ieee", "wos"],
        metavar="SOURCE",
        help="Sources to fetch (default: all three). Ignored with --corpus-only.",
    )
    p.add_argument("--max", type=int, default=config.MAX_RESULTS, dest="max_results")
    args, _ = p.parse_known_args()
    return args


# ---------------------------------------------------------------------------
# Substep 1 — fetch
# ---------------------------------------------------------------------------

def run_fetch(sources: list[str], max_results: int) -> None:
    """Fetch raw metadata from each requested source and save *_raw.csv."""
    ensure_output_dir("outputs/metadata")

    if "pubmed" in sources:
        ncbi_key   = os.environ.get("NCBI_API_KEY", "")
        ncbi_email = os.environ.get("NCBI_EMAIL", "")
        try:
            fetcher = PubMedFetcher(email=ncbi_email, api_key=ncbi_key)
            df = fetcher.fetch(max_results=max_results)
            df.to_csv(config.PUBMED_RAW_CSV, index=False)
            PubMedFetcher.write_clean_bib(df, config.PUBMED_BIB_OUT)
            logger.info("PubMed: %d records → %s", len(df), config.PUBMED_RAW_CSV)
        except Exception as exc:
            logger.error("PubMed fetch failed: %s", exc)

    if "ieee" in sources:
        try:
            loader = IEEEExportLoader(
                csv_path=config.IEEE_EXPORT_CSV,
                bib_path=config.IEEE_EXPORT_BIB,
            )
            df = loader.load()
            loader.write_clean_bib(config.IEEE_BIB_OUT)
            df.to_csv(config.IEEE_RAW_CSV, index=False)
            logger.info("IEEE: %d records → %s", len(df), config.IEEE_RAW_CSV)
        except Exception as exc:
            logger.error("IEEE load failed: %s", exc)

    if "wos" in sources:
        try:
            loader = WoSExportLoader(ris_path=config.WOS_EXPORT_RIS)
            df = loader.load()
            loader.write_clean_bib(config.WOS_BIB_OUT)
            df.to_csv(config.WOS_RAW_CSV, index=False)
            logger.info("WoS: %d records → %s", len(df), config.WOS_RAW_CSV)
        except Exception as exc:
            logger.error("WoS load failed: %s", exc)


# ---------------------------------------------------------------------------
# Substep 2 — corpus
# ---------------------------------------------------------------------------

def run_corpus() -> None:
    """Merge available *_raw.csv files, deduplicate, and write corpus outputs."""
    ensure_output_dir("outputs/metadata")

    raw_files = {
        "pubmed": config.PUBMED_RAW_CSV,
        "ieee":   config.IEEE_RAW_CSV,
        "wos":    config.WOS_RAW_CSV,
    }

    frames: list[pd.DataFrame] = []
    for source, path in raw_files.items():
        if Path(path).exists():
            df = pd.read_csv(path, dtype=str).fillna("")
            frames.append(df)
            logger.info("Loaded %d records from %s", len(df), path)
        else:
            logger.warning("Raw file not found, skipping: %s", path)

    if not frames:
        logger.error("No raw files found in outputs/metadata/. Run with --fetch-only first.")
        return

    combined = pd.concat(frames, ignore_index=True)

    # Assign a permanent sequential corpus_id to every row before dedup
    combined["corpus_id"] = list(range(1, len(combined) + 1))

    corpus = deduplicate_corpus(combined)

    # Enforce canonical column order from config
    corpus = corpus[config.CORPUS_COLUMNS]

    corpus.to_csv(config.CORPUS_CSV, index=False)
    corpus.to_excel(config.CORPUS_XLSX, index=False)
    joblib.dump(corpus, config.CORPUS_PKL)

    n_unique = (~corpus["is_duplicate"]).sum()
    logger.info(
        "Corpus: %d total rows, %d duplicates, %d unique records.",
        len(corpus), len(corpus) - n_unique, n_unique,
    )
    logger.info("Saved → %s | %s | %s", config.CORPUS_CSV, config.CORPUS_XLSX, config.CORPUS_PKL)

    dup_map = build_duplicate_map(corpus)
    dup_map_path = Path(config.CORPUS_CSV).parent / "duplicate_map.csv"
    dup_map.to_csv(dup_map_path, index=False)
    logger.info("Duplicate map: %d entries → %s", len(dup_map), dup_map_path)

    _write_screening_files(corpus)


def _write_screening_files(corpus: pd.DataFrame) -> None:
    """Write per-reviewer screening files (CSV + XLSX) to outputs/screening/."""
    ensure_output_dir("outputs/screening")

    unique = corpus[~corpus["is_duplicate"]][["corpus_id", "title", "abstract"]].copy().reset_index(drop=True)
    unique["phase"] = "1"
    unique["PO1_population_clinician"] = ""
    unique["CO2_concept_llm_presence"] = ""
    unique["CO3_concept_cdss_function"] = ""
    unique["CX4_context_clinical_data"] = ""
    unique["OT5_other"] = ""
    unique["decision"] = ""
    unique["rationale"] = ""

    col_widths = {
        "corpus_id": 10, "title": 55, "abstract": 80, "phase": 8,
        "PO1_population_clinician": 22, "CO2_concept_llm_presence": 22,
        "CO3_concept_cdss_function": 22, "CX4_context_clinical_data": 22,
        "OT5_other": 14, "decision": 12, "rationale": 45,
    }
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    for reviewer in ["R1", "R2"]:
        csv_path  = Path(f"outputs/screening/screening_phase1_{reviewer}.csv")
        xlsx_path = Path(f"outputs/screening/screening_phase1_{reviewer}.xlsx")

        unique.to_csv(csv_path, index=False)

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            unique.to_excel(writer, index=False, sheet_name="Phase1")
            ws = writer.sheets["Phase1"]
            for i, col in enumerate(unique.columns, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 15)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        logger.info("Screening files written → %s | %s", csv_path, xlsx_path)

    logger.info("Screening files: %d unique records, 2 reviewers.", len(unique))


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def main() -> None:
    setup_logging()
    args = parse_args()

    if not args.corpus_only:
        logger.info("=== Substep 1: fetch raw metadata ===")
        run_fetch(sources=args.sources, max_results=args.max_results)

    if not args.fetch_only:
        logger.info("=== Substep 2: build corpus ===")
        run_corpus()


if __name__ == "__main__":
    # Jupyter / interactive use — uncomment one:
    # sys.argv = ["run_metadata.py"]
    # sys.argv = ["run_metadata.py", "--fetch-only", "--sources", "pubmed"]
    # sys.argv = ["run_metadata.py", "--corpus-only"]
    main()
