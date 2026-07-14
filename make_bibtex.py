"""Export the standardized naming map and BibTeX file for Phase 1 includes.

For every included record the canonical name (see naming.py) is used as the
BibTeX citation key, identical to the PDF filename stem in papers_library:

    \\cite{Sandmann2025_45_prompt-design-clinical}  <->  Sandmann2025_45_prompt-design-clinical.pdf

Outputs:
    outputs/fulltext/naming_map.csv   corpus_id -> standard_name (+ title, doi)
    outputs/fulltext/includes.bib     one entry per include, standardized keys
                                      (source-provided BibTeX when available,
                                      minimal synthesized entry otherwise)

Usage:
    python make_bibtex.py
"""

import logging
import re

import pandas as pd
from openpyxl import load_workbook

import config
from naming import standard_name
from utils import setup_logging, ensure_output_dir

logger = logging.getLogger(__name__)


def load_worklist() -> pd.DataFrame:
    """Parse the Phase 1 compiled screening workbook and join includes with the corpus."""
    wb = load_workbook(config.PHASE1_APPEND_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    # Guard against column drift: row 2 holds the real column names
    if ws["A2"].value != "corpus_id" or ws["W2"].value != "Final decision":
        raise ValueError(
            f"Unexpected header in {config.PHASE1_APPEND_XLSX}: "
            f"A2={ws['A2'].value!r}, W2={ws['W2'].value!r} "
            "(expected 'corpus_id' / 'Final decision')."
        )

    decisions = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        corpus_id, final_decision = row[0], row[22]  # cols A and W
        if corpus_id is None:
            continue
        decisions.append({"corpus_id": int(corpus_id), "final_decision": final_decision})
    wb.close()

    counts = pd.Series([d["final_decision"] for d in decisions]).value_counts().to_dict()
    missing = [d["corpus_id"] for d in decisions if d["final_decision"] in (None, "")]
    if missing:
        raise ValueError(
            f"Final decision empty for corpus_ids {missing} — if the workbook was "
            "re-saved without cached formula values, open and re-save it in Excel/LibreOffice."
        )
    expected = config.PHASE1_EXPECTED
    if (counts.get("INCLUDE") != expected["INCLUDE"]
            or counts.get("EXCLUDE") != expected["EXCLUDE"]
            or len(decisions) != expected["TOTAL"]):
        raise ValueError(
            f"Decision counts {counts} (total {len(decisions)}) do not match "
            f"expected {expected}. Update PHASE1_EXPECTED in config.py if the "
            "screening file legitimately changed."
        )

    decisions_df = pd.DataFrame(decisions)
    ensure_output_dir("outputs/screening")
    decisions_df.to_csv(config.PHASE1_DECISIONS_CSV, index=False)
    logger.info("Phase 1 decisions: %s → %s", counts, config.PHASE1_DECISIONS_CSV)

    corpus = pd.read_csv(config.CORPUS_CSV, dtype=str).fillna("")
    corpus["corpus_id"] = corpus["corpus_id"].astype(int)
    includes = decisions_df[decisions_df["final_decision"] == "INCLUDE"]
    worklist = includes.merge(
        corpus[["corpus_id", "source", "uid", "title", "journal", "year", "authors", "doi", "url"]],
        on="corpus_id", how="inner",
    )
    if len(worklist) != expected["INCLUDE"] or worklist["corpus_id"].duplicated().any():
        raise ValueError(
            f"Worklist has {len(worklist)} rows after corpus join (expected "
            f"{expected['INCLUDE']}, no duplicates) — corpus.csv and the screening "
            "workbook are out of sync."
        )

    worklist["doi"] = worklist["doi"].apply(lambda d: _normalize_doi(d))
    worklist["pmid"] = worklist.apply(
        lambda r: r["uid"] if r["source"] == "pubmed" else "", axis=1
    )
    return worklist


def _normalize_doi(doi: str) -> str:
    doi = doi.strip().lower()
    return re.sub(r"^https?://(dx\.)?doi\.org/", "", doi)

logger = logging.getLogger(__name__)


def _rekey(bibtex: str, key: str) -> str:
    """Replace the citation key of a BibTeX entry."""
    return re.sub(r"@(\w+)\s*\{[^,]*,", rf"@\1{{{key},", bibtex, count=1)


def _synthesize(rec: pd.Series, key: str) -> str:
    """Minimal entry for records whose source export carried no BibTeX."""
    authors = " and ".join(a.strip() for a in rec["authors"].split(";") if a.strip())
    fields = [
        ("title", rec["title"]), ("author", authors), ("journal", rec["journal"]),
        ("year", rec["year"]), ("doi", rec["doi"]),
    ]
    body = ",\n".join(f"  {name} = {{{value}}}" for name, value in fields if value)
    return f"@article{{{key},\n{body}\n}}"


def main() -> None:
    setup_logging()
    worklist = load_worklist()
    corpus = pd.read_csv(config.CORPUS_CSV, dtype=str).fillna("")
    corpus["corpus_id"] = corpus["corpus_id"].astype(int)
    worklist = worklist.merge(corpus[["corpus_id", "bibtex"]], on="corpus_id")

    worklist["standard_name"] = worklist.apply(
        lambda r: standard_name(r["corpus_id"], r["authors"], r["year"], r["title"]), axis=1
    )
    dups = worklist[worklist["standard_name"].duplicated(keep=False)]
    if not dups.empty:
        raise ValueError(f"Duplicate standard names:\n{dups[['corpus_id', 'standard_name']]}")

    ensure_output_dir("outputs/fulltext")
    worklist[["corpus_id", "standard_name", "title", "doi"]].to_csv(
        config.NAMING_MAP_CSV, index=False
    )

    entries, synthesized = [], 0
    for _, rec in worklist.iterrows():
        if rec["bibtex"].strip():
            entries.append(_rekey(rec["bibtex"].strip(), rec["standard_name"]))
        else:
            entries.append(_synthesize(rec, rec["standard_name"]))
            synthesized += 1
    with open(config.INCLUDES_BIB, "w", encoding="utf-8") as fh:
        fh.write("\n\n".join(entries) + "\n")

    logger.info("Naming map: %d records → %s", len(worklist), config.NAMING_MAP_CSV)
    logger.info("BibTeX: %d entries (%d synthesized) → %s",
                len(entries), synthesized, config.INCLUDES_BIB)


if __name__ == "__main__":
    main()
