"""Stage 2b — full-text retrieval for Phase 1 included records.

Downloads full-text PDFs for the records marked INCLUDE in the Phase 1
title/abstract screening (screening/screening_phase1_append.xlsx), trying
open-access sources only:

  1. PMC Open Access subset  (PMCID resolved via the NCBI ID Converter)
  2. Unpaywall               (any OA location with a PDF, by DOI)

Everything that cannot be fetched automatically is written to a manual
worklist (outputs/fulltext/manual_worklist.xlsx) for hand retrieval.
PDFs land in papers_library/fulltext_retrieved/ named
{corpus_id}_{FirstAuthorSurname}_{year}.pdf.

Usage:
    python run_fulltext_retrieval.py                 # full run over all includes
    python run_fulltext_retrieval.py --only 5 12 40  # subset (smoke testing)
    python run_fulltext_retrieval.py --rescan        # no network: fold manually
                                                     # downloaded PDFs back into the log

Options:
    --only ID [ID ...]   Restrict retrieval to the given corpus_ids.
    --rescan             Re-walk the library folders for PDFs added or moved by
                         hand, update retrieval_log.csv and the manual worklist.
"""

import argparse
import logging
import os
import re
import tarfile
import time
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from tqdm import tqdm

import config
from naming import standard_name
from utils import setup_logging, ensure_output_dir, retry

logger = logging.getLogger(__name__)

IDCONV_URL    = "https://pmc.ncbi.nlm.nih.gov/tools/idconv/api/v1/articles/"
PMC_OA_URL    = "https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi"
UNPAYWALL_URL = "https://api.unpaywall.org/v2/{doi}"

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) lmm-cdss-review-pipeline/0.1"

LOG_COLUMNS = [
    "corpus_id", "standard_name", "title", "journal", "year", "source", "doi",
    "pmid", "pmcid", "status", "method", "attempts", "file_path",
    "manual_urls", "notes",
]


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Stage 2b: full-text retrieval for Phase 1 includes")
    p.add_argument("--only", nargs="+", type=int, metavar="ID", default=None,
                   help="Restrict to the given corpus_ids (smoke testing).")
    p.add_argument("--rescan", action="store_true",
                   help="No network: fold manually added/moved PDFs back into the log.")
    args, _ = p.parse_known_args()
    return args


# ---------------------------------------------------------------------------
# Substep 0 — Phase 1 decisions → retrieval worklist
# ---------------------------------------------------------------------------

def load_worklist() -> pd.DataFrame:
    """Parse the Phase 1 compiled screening workbook and join includes with the corpus."""
    wb = load_workbook(config.PHASE1_APPEND_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    # Guard against column drift: row 2 holds the real column names
    if ws["A2"].value != "corpus_id" or ws["W2"].value != "Final decision":
        raise ValueError(
            f"Unexpected header in {config.PHASE1_APPEND_XLSX}: "
            f"A2={ws['A2'].value!r}, W2={ws['W2'].value!r} "
            "(expected 'corpus_id' / 'Final decision')."
        )

    decisions = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        corpus_id, final_decision = row[0], row[22]  # cols A and W
        if corpus_id is None:
            continue
        decisions.append({"corpus_id": int(corpus_id), "final_decision": final_decision})
    wb.close()

    counts = pd.Series([d["final_decision"] for d in decisions]).value_counts().to_dict()
    missing = [d["corpus_id"] for d in decisions if d["final_decision"] in (None, "")]
    if missing:
        raise ValueError(
            f"Final decision empty for corpus_ids {missing} — if the workbook was "
            "re-saved without cached formula values, open and re-save it in Excel/LibreOffice."
        )
    expected = config.PHASE1_EXPECTED
    if (counts.get("INCLUDE") != expected["INCLUDE"]
            or counts.get("EXCLUDE") != expected["EXCLUDE"]
            or len(decisions) != expected["TOTAL"]):
        raise ValueError(
            f"Decision counts {counts} (total {len(decisions)}) do not match "
            f"expected {expected}. Update PHASE1_EXPECTED in config.py if the "
            "screening file legitimately changed."
        )

    decisions_df = pd.DataFrame(decisions)
    ensure_output_dir("outputs/screening")
    decisions_df.to_csv(config.PHASE1_DECISIONS_CSV, index=False)
    logger.info("Phase 1 decisions: %s → %s", counts, config.PHASE1_DECISIONS_CSV)

    corpus = pd.read_csv(config.CORPUS_CSV, dtype=str).fillna("")
    corpus["corpus_id"] = corpus["corpus_id"].astype(int)
    includes = decisions_df[decisions_df["final_decision"] == "INCLUDE"]
    worklist = includes.merge(
        corpus[["corpus_id", "source", "uid", "title", "journal", "year", "authors", "doi", "url"]],
        on="corpus_id", how="inner",
    )
    if len(worklist) != expected["INCLUDE"] or worklist["corpus_id"].duplicated().any():
        raise ValueError(
            f"Worklist has {len(worklist)} rows after corpus join (expected "
            f"{expected['INCLUDE']}, no duplicates) — corpus.csv and the screening "
            "workbook are out of sync."
        )

    worklist["doi"] = worklist["doi"].map(_normalize_doi)
    worklist["pmid"] = worklist.apply(
        lambda r: r["uid"] if r["source"] == "pubmed" else "", axis=1
    )
    return worklist


def _normalize_doi(doi: str) -> str:
    doi = doi.strip().lower()
    return re.sub(r"^https?://(dx\.)?doi\.org/", "", doi)


# ---------------------------------------------------------------------------
# Substep 1 — PMCID resolution (NCBI ID Converter, one batch call)
# ---------------------------------------------------------------------------

def resolve_pmcids(worklist: pd.DataFrame, session: requests.Session, email: str) -> dict[int, str]:
    """Map corpus_id → PMCID via the NCBI ID Converter (one batch per id type)."""
    pmids = {rec["pmid"]: rec["corpus_id"]
             for _, rec in worklist.iterrows() if rec["pmid"]}
    dois = {rec["doi"]: rec["corpus_id"]
            for _, rec in worklist.iterrows() if not rec["pmid"] and rec["doi"]}

    pmcids: dict[int, str] = {}
    for idtype, id_to_corpus in (("pmid", pmids), ("doi", dois)):
        ids = list(id_to_corpus)
        for i in range(0, len(ids), 150):  # API limit: 200 ids per request
            records = _idconv_batch(session, email, ids[i:i + 150], idtype)
            for record in records:
                requested = str(record.get("requested-id", ""))
                if record.get("pmcid") and requested in id_to_corpus:
                    pmcids[id_to_corpus[requested]] = record["pmcid"]
    return pmcids


@retry(max_attempts=3, wait=5.0)
def _idconv_batch(session: requests.Session, email: str,
                  ids: list[str], idtype: str) -> list[dict]:
    resp = _polite_get(session, IDCONV_URL, params={
        "tool": "lmm-cdss-review-pipeline", "email": email,
        "ids": ",".join(ids), "idtype": idtype, "format": "json", "versions": "no",
    })
    return resp.json().get("records", [])


# ---------------------------------------------------------------------------
# Retrieval sources
# ---------------------------------------------------------------------------

def _polite_get(session: requests.Session, url: str, **kwargs) -> requests.Response:
    time.sleep(config.FULLTEXT_SLEEP)
    resp = session.get(url, timeout=config.FULLTEXT_TIMEOUT, **kwargs)
    resp.raise_for_status()
    return resp


def pmc_oa_download(session: requests.Session, pmcid: str, dest: Path) -> tuple[bool, str]:
    """Fetch the article PDF from the PMC Open Access service.

    Most OA-subset records only expose a .tar.gz package (no direct PDF link);
    the package contains the article PDF, which is extracted. The ftp host is
    served over https and is never bot-blocked. Returns (success, attempt_note).
    """
    resp = _polite_get(session, PMC_OA_URL, params={"id": pmcid})
    root = ET.fromstring(resp.content)
    if root.find(".//error") is not None:
        return False, "not_open_access"

    def _https(link) -> str:
        href = link.get("href", "").replace(
            "ftp://ftp.ncbi.nlm.nih.gov", "https://ftp.ncbi.nlm.nih.gov"
        )
        # oa.fcgi still hands out legacy /pub/pmc/oa_package/ hrefs; the files
        # moved to /pub/pmc/deprecated/ and stay there until Aug 2026 (NCBI
        # readme 2026-04-10) — the AWS replacement bucket carries no PDFs.
        return href.replace("/pub/pmc/oa_package/", "/pub/pmc/deprecated/oa_package/")

    pdf_link = root.find(".//link[@format='pdf']")
    if pdf_link is not None:
        if download_pdf(session, _https(pdf_link), dest):
            return True, "pdf"
        return False, "invalid_pdf"

    tgz_link = root.find(".//link[@format='tgz']")
    if tgz_link is None:
        return False, "no_package"
    return _extract_pdf_from_tgz(session, _https(tgz_link), dest)


def _extract_pdf_from_tgz(session: requests.Session, url: str, dest: Path) -> tuple[bool, str]:
    """Download an OA package and extract its (largest) PDF member to *dest*."""
    tgz_part = dest.with_suffix(".tgz.part")
    try:
        with _polite_get(session, url, stream=True) as resp, open(tgz_part, "wb") as fh:
            for chunk in resp.iter_content(chunk_size=65536):
                fh.write(chunk)
        with tarfile.open(tgz_part, "r:gz") as tar:
            members = [m for m in tar.getmembers()
                       if m.isfile() and m.name.lower().endswith(".pdf")]
            if not members:
                return False, "no_pdf_in_package"
            member = max(members, key=lambda m: m.size)
            content = tar.extractfile(member).read()
        if b"%PDF" not in content[:1024]:
            return False, "invalid_pdf"
        part = dest.with_suffix(dest.suffix + ".part")
        part.write_bytes(content)
        part.rename(dest)
        return True, "tgz"
    finally:
        tgz_part.unlink(missing_ok=True)


def unpaywall_pdf_urls(session: requests.Session, doi: str, email: str) -> list[str]:
    """Candidate PDF URLs from Unpaywall, best OA location first."""
    try:
        resp = _polite_get(session, UNPAYWALL_URL.format(doi=doi), params={"email": email})
    except requests.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return []  # DOI unknown to Unpaywall
        raise
    data = resp.json()
    locations = [data.get("best_oa_location") or {}] + (data.get("oa_locations") or [])
    urls = [loc.get("url_for_pdf") for loc in locations if loc.get("url_for_pdf")]
    return list(dict.fromkeys(urls))


def download_pdf(session: requests.Session, url: str, dest: Path) -> bool:
    """Stream *url* to *dest*; reject anything without PDF magic bytes."""
    part = dest.with_suffix(dest.suffix + ".part")
    try:
        with _polite_get(session, url, stream=True) as resp:
            chunks = resp.iter_content(chunk_size=65536)
            first = next(chunks, b"")
            if b"%PDF" not in first[:1024]:
                return False  # HTML error page, not a PDF
            with open(part, "wb") as fh:
                fh.write(first)
                for chunk in chunks:
                    fh.write(chunk)
        part.rename(dest)
        return True
    except Exception:
        part.unlink(missing_ok=True)
        raise


# ---------------------------------------------------------------------------
# Substep 2 — per-record cascade
# ---------------------------------------------------------------------------

def make_filename(rec: pd.Series) -> str:
    return standard_name(rec["corpus_id"], rec["authors"], rec["year"], rec["title"]) + ".pdf"


def find_existing(corpus_id: int) -> Path | None:
    """Locate an already-present PDF for this record across the library lifecycle folders.

    Matches the standard name (corpus_id is the only all-digit field between
    underscores: Surname2025_{id}_title.pdf) plus legacy/lazy patterns
    ({id}_*.pdf, {id}.pdf) so manually dropped files are still recognized.
    """
    library = Path(config.FULLTEXT_DIR).parent
    for folder in ("fulltext_retrieved", "included", "excluded"):
        for pattern in (f"*_{corpus_id}_*.pdf", f"{corpus_id}_*.pdf", f"{corpus_id}.pdf"):
            hits = sorted((library / folder).glob(pattern))
            if hits:
                return hits[0]
    return None


def retrieve_one(rec: pd.Series, pmcid: str, session: requests.Session,
                 email: str) -> dict:
    dest = Path(config.FULLTEXT_DIR) / make_filename(rec)
    result = {
        "corpus_id": rec["corpus_id"], "standard_name": dest.stem,
        "title": rec["title"], "year": rec["year"],
        "source": rec["source"], "doi": rec["doi"], "pmid": rec["pmid"], "pmcid": pmcid,
        "status": "manual_needed", "method": "", "attempts": [],
        "file_path": "", "manual_urls": "", "notes": "",
    }

    existing = find_existing(rec["corpus_id"])
    if existing:
        result.update(status="retrieved", method="existing_file", file_path=str(existing))
        return result

    # 1. PMC Open Access subset
    if pmcid:
        try:
            ok, note = pmc_oa_download(session, pmcid, dest)
            if ok:
                result.update(status="retrieved", method=f"pmc_oa_{note}", file_path=str(dest))
                return result
            result["attempts"].append(f"pmc_oa:{note}")
        except Exception as exc:
            result["attempts"].append(f"pmc_oa:error:{type(exc).__name__}")
    else:
        result["attempts"].append("pmc_oa:no_pmcid")

    # 2. Unpaywall OA locations
    upw_urls: list[str] = []
    if rec["doi"]:
        try:
            upw_urls = unpaywall_pdf_urls(session, rec["doi"], email)
            if not upw_urls:
                result["attempts"].append("unpaywall:no_oa_pdf")
            for url in upw_urls:
                try:
                    if download_pdf(session, url, dest):
                        result.update(status="retrieved", method="unpaywall", file_path=str(dest))
                        return result
                    result["attempts"].append("unpaywall:invalid_pdf")
                except Exception as exc:
                    result["attempts"].append(f"unpaywall:error:{type(exc).__name__}")
        except Exception as exc:
            result["attempts"].append(f"unpaywall:error:{type(exc).__name__}")
    else:
        result["attempts"].append("unpaywall:no_doi")

    # 3. Manual worklist — best links first: the PMC article page and the OA
    # PDF urls Unpaywall offered usually work in a real browser even though
    # scripted downloads get bot-blocked (403).
    urls = []
    if pmcid:
        urls.append(f"https://pmc.ncbi.nlm.nih.gov/articles/{pmcid}/")
    urls.extend(upw_urls)
    if rec["doi"]:
        urls.append(f"https://doi.org/{rec['doi']}")
    if rec["url"]:
        urls.append(rec["url"])
    result["manual_urls"] = " ; ".join(dict.fromkeys(urls))
    return result


# ---------------------------------------------------------------------------
# Substep 3 — outputs
# ---------------------------------------------------------------------------

def write_outputs(results: pd.DataFrame) -> None:
    ensure_output_dir("outputs/fulltext")
    results = results[LOG_COLUMNS]
    results.to_csv(config.FULLTEXT_LOG_CSV, index=False)

    manual = results[results["status"] == "manual_needed"][
        ["corpus_id", "standard_name", "title", "year", "journal", "doi",
         "manual_urls", "notes"]
    ].copy()
    manual["resolved"] = ""

    col_widths = {"corpus_id": 10, "standard_name": 38, "title": 60, "year": 8,
                  "journal": 30, "doi": 32, "manual_urls": 60, "notes": 40,
                  "resolved": 10}
    with pd.ExcelWriter(config.FULLTEXT_MANUAL_XLSX, engine="openpyxl") as writer:
        manual.to_excel(writer, index=False, sheet_name="Manual")
        ws = writer.sheets["Manual"]
        for i, col in enumerate(manual.columns, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 15)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    n_ok = (results["status"] == "retrieved").sum()
    logger.info("Full text: %d retrieved, %d manual (of %d).", n_ok, len(manual), len(results))
    logger.info("Saved → %s | %s", config.FULLTEXT_LOG_CSV, config.FULLTEXT_MANUAL_XLSX)


def run_rescan() -> None:
    """Fold manually downloaded/moved PDFs back into the retrieval log (no network)."""
    if not Path(config.FULLTEXT_LOG_CSV).exists():
        logger.error("No %s found — run the retrieval first.", config.FULLTEXT_LOG_CSV)
        return
    results = pd.read_csv(config.FULLTEXT_LOG_CSV, dtype=str).fillna("")
    results["corpus_id"] = results["corpus_id"].astype(int)

    flipped = 0
    for idx, row in results.iterrows():
        existing = find_existing(row["corpus_id"])
        if existing and row["status"] != "retrieved":
            results.loc[idx, ["status", "method", "file_path"]] = \
                ["retrieved", "manual_added", str(existing)]
            flipped += 1
        elif row["status"] == "retrieved" and row["method"] != "manual_added" and not existing:
            logger.warning("corpus_id %s marked retrieved but file is gone: %s",
                           row["corpus_id"], row["file_path"])
    logger.info("Rescan: %d record(s) newly resolved by hand.", flipped)
    write_outputs(results)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def main() -> None:
    setup_logging()
    args = parse_args()

    if args.rescan:
        logger.info("=== Rescan: fold manually added PDFs into the log ===")
        run_rescan()
        return

    email = os.environ.get("NCBI_EMAIL", "")
    if not email:
        raise ValueError("NCBI_EMAIL is required but not set in .env")
    upw_email = os.environ.get("UNPAYWALL_EMAIL", email)

    logger.info("=== Substep 0: Phase 1 decisions → worklist ===")
    worklist = load_worklist()
    if args.only:
        worklist = worklist[worklist["corpus_id"].isin(args.only)]
        logger.info("Restricted to %d record(s): %s", len(worklist), args.only)
    ensure_output_dir(config.FULLTEXT_DIR)

    session = requests.Session()
    session.headers["User-Agent"] = USER_AGENT

    logger.info("=== Substep 1: resolve PMCIDs (NCBI ID Converter) ===")
    pmcids = resolve_pmcids(worklist, session, email)
    logger.info("PMCIDs found for %d of %d records.", len(pmcids), len(worklist))

    logger.info("=== Substep 2: retrieve full texts ===")
    rows = []
    for _, rec in tqdm(worklist.iterrows(), total=len(worklist), desc="Full text"):
        try:
            result = retrieve_one(rec, pmcids.get(rec["corpus_id"], ""), session, upw_email)
        except Exception as exc:
            logger.error("corpus_id %s failed: %s", rec["corpus_id"], exc)
            result = {"corpus_id": rec["corpus_id"],
                      "standard_name": Path(make_filename(rec)).stem,
                      "title": rec["title"],
                      "year": rec["year"], "source": rec["source"], "doi": rec["doi"],
                      "pmid": rec["pmid"], "pmcid": pmcids.get(rec["corpus_id"], ""),
                      "status": "manual_needed", "method": "",
                      "attempts": [f"error:{type(exc).__name__}"], "file_path": "",
                      "manual_urls": f"https://doi.org/{rec['doi']}" if rec["doi"] else rec["url"],
                      "notes": str(exc)}
        result["journal"] = rec["journal"]
        result["attempts"] = "; ".join(result["attempts"])
        rows.append(result)

    logger.info("=== Substep 3: write retrieval log + manual worklist ===")
    write_outputs(pd.DataFrame(rows))


if __name__ == "__main__":
    # Jupyter / interactive use — uncomment one:
    # sys.argv = ["run_fulltext_retrieval.py"]
    # sys.argv = ["run_fulltext_retrieval.py", "--only", "5", "12"]
    main()
